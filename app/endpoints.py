import os
import zipfile
import sqlite3
import csv
import io
import json
import logging
import re
from pathlib import Path
from datetime import datetime, timezone
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi import (
    APIRouter,
    HTTPException,
    Request,
    Depends,
    Response,
    File,
    UploadFile,
    BackgroundTasks,
)
from fastapi.responses import StreamingResponse
from typing import List, Optional
import qrcode

from .models import (
    Item,
    ItemCreate,
    ItemUpdate,
    Settings,
    AnalyticsData,
    Subscriber,
    Message,
    SubscribeRequest,
    ContactRequest,
    ReorderRequest,
    ImageUploadResponse,
    Page,
    PageCreate,
    PageUpdate,
)
from .database import (
    create_item_in_db,
    update_item_in_db,
    delete_item_from_db,
    get_next_display_order,
    get_db_connection,
    get_settings_from_db,
    get_page_by_slug,
    get_page_by_id,
    get_all_pages,
    create_page,
    update_page,
    delete_page,
    get_mediakit_data,
    update_mediakit_data,
    delete_mediakit_entry,
    save_social_stats_cache,
    get_social_stats_cache,
    get_mediakit_setting,
    update_mediakit_setting,
    get_all_mediakit_settings,
    track_mediakit_view,
    get_mediakit_views,
    get_mediakit_views_stats,
    create_access_request,
    get_access_requests,
    update_access_request_status,
    get_mediakit_blocks,
    get_visible_mediakit_blocks,
    create_mediakit_block,
    update_mediakit_block,
    delete_mediakit_block,
    reorder_mediakit_blocks,
    check_access_approved,
)
from .auth_unified import require_auth, check_auth
from .services import (
    scrape_link_details,
    get_video_embed_url,
    save_optimized_image,
    generate_social_card,
    get_country_from_ip,
    APP_DOMAIN,
)
from .social_stats import get_stats_service
from .rate_limit import limiter_strict, limiter_standard
from .cache_unified import cache
from .config import BASE_DIR, UPLOAD_DIR

router = APIRouter()

# Slug validation pattern
SLUG_PATTERN = re.compile(r'^[a-z0-9-]+$')

# --- Background Task for Scraping ---


async def background_scrape_and_update(item_id: int, url: str):
    """Background task to scrape link details and update the database entry."""
    try:
        details = await scrape_link_details(url)
        update_data = {"title": details.get("title", "?"), "image_url": details.get("image_url")}
        # Only update if we got meaningful data
        if update_data["title"] != "?":
            update_item_in_db(item_id, update_data)
            cache.invalidate("items")
    except Exception as e:
        # Log error but don't crash - the item already exists with placeholder data
        logging.error(f"Background scraping failed for item {item_id}: {e}")


# --- Auth Check ---


@router.get("/auth/check", dependencies=[Depends(limiter_strict)])
async def check_login(username: str = Depends(require_auth)):
    return {"status": "ok"}


# --- Page Management ---


@router.get("/pages", response_model=List[Page])
async def get_pages(user=Depends(require_auth)):
    """Get all pages."""
    pages = get_all_pages()
    return [Page(**page) for page in pages]


@router.get("/pages/{page_id}", response_model=Page)
async def get_page(page_id: int, user=Depends(require_auth)):
    """Get a specific page by ID."""
    page = get_page_by_id(page_id)
    if not page:
        raise HTTPException(404, "Page nicht gefunden")
    return Page(**page)


@router.post("/pages", response_model=Page)
async def create_new_page(page_data: PageCreate, user=Depends(require_auth)):
    """Create a new page."""
    # Validate slug format
    if not SLUG_PATTERN.match(page_data.slug):
        raise HTTPException(400, "Slug darf nur Kleinbuchstaben, Zahlen und Bindestriche enthalten")
    
    # Check if slug already exists
    existing = get_page_by_slug(page_data.slug)
    if existing:
        raise HTTPException(400, "Eine Seite mit diesem Slug existiert bereits")
    page = create_page(
        slug=page_data.slug,
        title=page_data.title,
        bio=page_data.bio or "",
        image_url=page_data.image_url or "",
        bg_image_url=page_data.bg_image_url or "",
    )
    cache.invalidate("items")
    return Page(**page)


@router.put("/pages/{page_id}", response_model=Page)
async def update_existing_page(page_id: int, page_data: PageUpdate, user=Depends(require_auth)):
    """Update a page."""
    # If slug is being updated, validate and check it doesn't conflict
    if page_data.slug:
        if not SLUG_PATTERN.match(page_data.slug):
            raise HTTPException(400, "Slug darf nur Kleinbuchstaben, Zahlen und Bindestriche enthalten")
        existing = get_page_by_slug(page_data.slug)
        if existing and existing["id"] != page_id:
            raise HTTPException(400, "Eine Seite mit diesem Slug existiert bereits")
    
    updated = update_page(page_id, page_data.model_dump(exclude_unset=True))
    if not updated:
        raise HTTPException(404, "Page nicht gefunden")
    cache.invalidate("items")
    return Page(**updated)


@router.delete("/pages/{page_id}")
async def delete_existing_page(page_id: int, user=Depends(require_auth)):
    """Delete a page and all its items."""
    page = get_page_by_id(page_id)
    if not page:
        raise HTTPException(404, "Page nicht gefunden")
    # Don't allow deleting the default page (slug = '')
    if page["slug"] == "":
        raise HTTPException(400, "Die Hauptseite kann nicht gelöscht werden")
    delete_page(page_id)
    cache.invalidate("items")
    return Response(status_code=204)


# --- Helper ---


def build_item_data(item_type, title, url=None, image_url=None, price=None, grid_columns=2, page_id=None):
    # Reihenfolge muss exakt zur DB-Tabelle passen:
    # item_type, title, url, image_url, display_order, parent_id, click_count, is_featured, is_active, is_affiliate, publish_on, expires_on, price, grid_columns, page_id
    return (
        item_type,
        title,
        url,
        image_url,
        get_next_display_order(page_id),
        None,
        0,
        0,
        1,
        0,
        None,
        None,
        price,
        grid_columns,
        page_id,
    )


# --- Item Creation Endpoints ---


@router.post("/links", response_model=Item)
async def create_link(req: ItemCreate, background_tasks: BackgroundTasks, user=Depends(require_auth)):
    if not req.url:
        raise HTTPException(400, "URL fehlt")
    # Create item immediately with placeholder
    data = build_item_data("link", "Loading...", req.url, None, page_id=req.page_id)
    item_dict = create_item_in_db(data)
    # Schedule background scraping
    background_tasks.add_task(background_scrape_and_update, item_dict["id"], req.url)
    cache.invalidate("items")
    return Item(**item_dict)


@router.post("/videos", response_model=Item)
async def create_video(req: ItemCreate, user=Depends(require_auth)):
    if not req.url:
        raise HTTPException(400, "URL fehlt")
    embed = get_video_embed_url(req.url) or req.url
    data = build_item_data("video", "Video", embed, page_id=req.page_id)
    cache.invalidate("items")
    return Item(**create_item_in_db(data))


@router.post("/headers", response_model=Item)
async def create_header(req: ItemCreate, user=Depends(require_auth)):
    # Cache Invalidation hinzugefügt
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("header", req.title, page_id=req.page_id)))


@router.post("/slider_groups", response_model=Item)
async def create_slider(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("slider_group", req.title, page_id=req.page_id)))


@router.post("/grids", response_model=Item)
async def create_grid(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("grid", req.title, page_id=req.page_id)))


@router.post("/faqs", response_model=Item)
async def create_faq(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("faq", req.title, "", page_id=req.page_id)))


@router.post("/dividers", response_model=Item)
async def create_divider(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("divider", req.title or "---", page_id=req.page_id)))


@router.post("/testimonials", response_model=Item)
async def create_testimonial(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("testimonial", req.name, req.text, page_id=req.page_id)))


@router.post("/products", response_model=Item)
async def create_product(req: ItemCreate, background_tasks: BackgroundTasks, user=Depends(require_auth)):
    if not req.url:
        raise HTTPException(400, "URL fehlt")
    # Create item immediately with placeholder
    title = req.title if req.title else "Loading..."
    data = build_item_data("product", title, req.url, None, req.price, page_id=req.page_id)
    item_dict = create_item_in_db(data)
    # Schedule background scraping if no custom title was provided
    if not req.title:
        background_tasks.add_task(background_scrape_and_update, item_dict["id"], req.url)
    cache.invalidate("items")
    return Item(**item_dict)


@router.post("/contact_form", response_model=Item)
async def create_contact_form(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("contact_form", req.title, page_id=req.page_id)))


@router.post("/email_form", response_model=Item)
async def create_email_form(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("email_form", req.title, page_id=req.page_id)))


@router.post("/countdowns", response_model=Item)
async def create_countdown(req: ItemCreate, user=Depends(require_auth)):
    cache.invalidate("items")
    return Item(**create_item_in_db(build_item_data("countdown", req.title, req.target_datetime, page_id=req.page_id)))


# --- Item Management ---


@router.put("/items/{id}", response_model=Item)
async def update_item(id: int, item: ItemUpdate, user=Depends(require_auth)):
    updated = update_item_in_db(id, item.model_dump(exclude_unset=True))
    if not updated:
        raise HTTPException(404, "Item nicht gefunden")
    cache.invalidate("items")
    return Item(**updated)


@router.delete("/items/{id}")
async def delete_item(id: int, user=Depends(require_auth)):
    delete_item_from_db(id)
    cache.invalidate("items")
    return Response(status_code=204)


@router.put("/items/{id}/toggle_visibility", response_model=Item)
async def toggle_visibility(id: int, user=Depends(require_auth)):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE items SET is_active = NOT is_active WHERE id = ?", (id,))
        conn.commit()
        cursor.execute("SELECT * FROM items WHERE id = ?", (id,))
        updated = cursor.fetchone()
    if not updated:
        raise HTTPException(404, "Item nicht gefunden")
    cache.invalidate("items")
    return Item(**dict(updated))


@router.post("/items/reorder")
async def reorder(req: ReorderRequest, user=Depends(require_auth)):
    with get_db_connection() as conn:
        for idx, iid in enumerate(req.ids):
            conn.execute("UPDATE items SET display_order = ? WHERE id = ?", (idx, iid))
        conn.commit()
    cache.invalidate("items")
    return Response(status_code=204)


# --- Uploads & Media ---


@router.post("/upload_image", response_model=ImageUploadResponse)
async def upload_image(file: UploadFile = File(...), user=Depends(require_auth)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(400, "Ungültiger Dateityp.")
    try:
        url = save_optimized_image(file, UPLOAD_DIR)
        return ImageUploadResponse(url=url)
    except Exception as e:
        raise HTTPException(500, f"Upload Fehler: {e}")


@router.get("/media/files")
async def list_media_files(user=Depends(require_auth)):
    files = []
    if os.path.exists(UPLOAD_DIR):
        try:
            entries = sorted(os.scandir(UPLOAD_DIR), key=lambda e: e.stat().st_mtime, reverse=True)
            for entry in entries:
                if entry.is_file() and entry.name.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")):
                    files.append(
                        {"name": entry.name, "url": f"/static/uploads/{entry.name}", "size": entry.stat().st_size}
                    )
        except:
            pass
    return files


@router.delete("/media/files/{filename}")
async def delete_media_file(filename: str, user=Depends(require_auth)):
    safe_name = os.path.basename(filename)
    path = UPLOAD_DIR / safe_name
    if not path.exists():
        raise HTTPException(404, "Datei nicht gefunden")
    try:
        os.remove(path)
        return Response(status_code=204)
    except Exception as e:
        raise HTTPException(500, f"Löschfehler: {e}")


# --- Settings & Backup ---


@router.get("/settings", response_model=Settings, dependencies=[Depends(limiter_standard)])
async def get_settings():
    cached = cache.get("settings")
    if cached:
        return cached
    settings = Settings(**get_settings_from_db())
    cache.set("settings", settings, ttl=300)
    return settings


@router.put("/settings", response_model=Settings)
async def update_settings(settings: Settings, user=Depends(require_auth)):
    with get_db_connection() as conn:
        cursor = conn.cursor()
        for k, v in settings.model_dump(exclude_unset=True).items():
            if v is not None:
                cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, v))
        conn.commit()
    cache.invalidate("settings")
    return Settings(**get_settings_from_db())


@router.get("/backup/download")
async def download_backup(user=Depends(require_auth)):
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            if os.path.exists("linktree.db"):
                zf.write("linktree.db")
            if os.path.exists(UPLOAD_DIR):
                for root, _, files in os.walk(UPLOAD_DIR):
                    for f in files:
                        zf.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), UPLOAD_DIR.parent))
        buf.seek(0)
        filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
        return StreamingResponse(
            buf, media_type="application/zip", headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(500, f"Backup Fehler: {e}")


# --- Analytics & Community ---


@router.get("/analytics", response_model=AnalyticsData)
async def get_analytics(user=Depends(require_auth)):
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(id) FROM clicks")
        total_clicks = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(id) FROM subscribers")
        total_subs = cur.fetchone()[0] or 0

        cur.execute(
            "SELECT date(timestamp) as day, COUNT(id) as clicks FROM clicks WHERE timestamp >= date('now', '-30 days') GROUP BY day ORDER BY day ASC"
        )
        clicks_per_day = [dict(r) for r in cur.fetchall()]

        cur.execute(
            "SELECT i.id, i.title, COUNT(c.id) as clicks FROM clicks c JOIN items i ON c.item_id = i.id GROUP BY i.id, i.title ORDER BY clicks DESC LIMIT 10"
        )
        top_links = [dict(r) for r in cur.fetchall()]

        cur.execute(
            "SELECT CASE WHEN referer IS NULL OR referer = '' THEN '(Direkt)' ELSE referer END as referer_domain, COUNT(id) as clicks FROM clicks GROUP BY referer_domain ORDER BY clicks DESC LIMIT 10"
        )
        top_referers = [dict(r) for r in cur.fetchall()]

        cur.execute(
            "SELECT CASE WHEN country_code IS NULL THEN 'Unbekannt' ELSE country_code END as country, COUNT(id) as clicks FROM clicks GROUP BY country ORDER BY clicks DESC LIMIT 10"
        )
        top_countries = [dict(r) for r in cur.fetchall()]

        return AnalyticsData(
            total_clicks=total_clicks,
            clicks_per_day=clicks_per_day,
            top_links=top_links,
            top_referers=top_referers,
            top_countries=top_countries,
            total_subscribers=total_subs,
        )


@router.get("/analytics/advanced")
async def get_advanced_analytics(
    user=Depends(require_auth),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    item_id: Optional[int] = None,
    country: Optional[str] = None,
    referer: Optional[str] = None,
):
    """
    Advanced analytics endpoint with filtering capabilities.
    Supports filtering by date range, item, country, and referrer.
    """
    with get_db_connection() as conn:
        cur = conn.cursor()
        
        # Build WHERE clauses for filters
        where_clauses = []
        params = []
        
        if start_date:
            where_clauses.append("date(timestamp) >= ?")
            params.append(start_date)
        
        if end_date:
            where_clauses.append("date(timestamp) <= ?")
            params.append(end_date)
        
        if item_id:
            where_clauses.append("item_id = ?")
            params.append(item_id)
        
        if country and country != "all":
            if country == "unknown":
                where_clauses.append("(country_code IS NULL OR country_code = '')")
            else:
                where_clauses.append("country_code = ?")
                params.append(country)
        
        if referer and referer != "all":
            if referer == "direct":
                where_clauses.append("(referer IS NULL OR referer = '')")
            else:
                where_clauses.append("referer = ?")
                params.append(referer)
        
        where_clause = " AND ".join(where_clauses) if where_clauses else "1=1"
        
        # Total clicks with filters
        cur.execute(f"SELECT COUNT(id) FROM clicks WHERE {where_clause}", params)
        total_clicks = cur.fetchone()[0] or 0
        
        # Clicks per day (last 30 days by default, or filtered range)
        date_range_days = 30
        if start_date and end_date:
            # Calculate days between dates
            from datetime import datetime as dt
            try:
                d1 = dt.strptime(start_date, "%Y-%m-%d")
                d2 = dt.strptime(end_date, "%Y-%m-%d")
                date_range_days = max((d2 - d1).days, 1)
            except (ValueError, TypeError):
                date_range_days = 30
        
        cur.execute(
            f"SELECT date(timestamp) as day, COUNT(id) as clicks FROM clicks WHERE {where_clause} GROUP BY day ORDER BY day DESC LIMIT ?",
            params + [min(date_range_days, 365)]
        )
        clicks_per_day = [dict(r) for r in cur.fetchall()]
        clicks_per_day.reverse()  # Show chronologically
        
        # Top links with filters
        cur.execute(
            f"SELECT i.id, i.title, COUNT(c.id) as clicks FROM clicks c JOIN items i ON c.item_id = i.id WHERE {where_clause} GROUP BY i.id, i.title ORDER BY clicks DESC LIMIT 20",
            params
        )
        top_links = [dict(r) for r in cur.fetchall()]
        
        # Top referrers with filters
        cur.execute(
            f"SELECT CASE WHEN referer IS NULL OR referer = '' THEN '(Direkt)' ELSE referer END as referer_domain, COUNT(id) as clicks FROM clicks WHERE {where_clause} GROUP BY referer_domain ORDER BY clicks DESC LIMIT 20",
            params
        )
        top_referers = [dict(r) for r in cur.fetchall()]
        
        # Top countries with filters
        cur.execute(
            f"SELECT CASE WHEN country_code IS NULL THEN 'Unbekannt' ELSE country_code END as country, COUNT(id) as clicks FROM clicks WHERE {where_clause} GROUP BY country ORDER BY clicks DESC LIMIT 20",
            params
        )
        top_countries = [dict(r) for r in cur.fetchall()]
        
        # Hourly distribution
        cur.execute(
            f"SELECT CAST(strftime('%H', timestamp) AS INTEGER) as hour, COUNT(id) as clicks FROM clicks WHERE {where_clause} GROUP BY hour ORDER BY hour ASC",
            params
        )
        clicks_per_hour = [dict(r) for r in cur.fetchall()]
        
        # Clicks by day of week
        cur.execute(
            f"SELECT CAST(strftime('%w', timestamp) AS INTEGER) as day_of_week, COUNT(id) as clicks FROM clicks WHERE {where_clause} GROUP BY day_of_week ORDER BY day_of_week ASC",
            params
        )
        clicks_per_weekday = [dict(r) for r in cur.fetchall()]
        
        return {
            "total_clicks": total_clicks,
            "clicks_per_day": clicks_per_day,
            "top_links": top_links,
            "top_referers": top_referers,
            "top_countries": top_countries,
            "clicks_per_hour": clicks_per_hour,
            "clicks_per_weekday": clicks_per_weekday,
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "item_id": item_id,
                "country": country,
                "referer": referer,
            }
        }


@router.get("/subscribers", response_model=List[Subscriber])
async def get_subscribers(user=Depends(require_auth)):
    with get_db_connection() as conn:
        conn.cursor().execute("SELECT id, email, subscribed_at FROM subscribers ORDER BY subscribed_at DESC")
        return [dict(r) for r in conn.cursor().fetchall()]


@router.delete("/subscribers/{id}")
async def delete_subscriber(id: int, user=Depends(require_auth)):
    with get_db_connection() as conn:
        conn.execute("DELETE FROM subscribers WHERE id = ?", (id,))
        conn.commit()
    return Response(status_code=204)


@router.get("/subscribers/export")
async def export_subscribers(user=Depends(require_auth)):
    with get_db_connection() as conn:
        subs = conn.execute("SELECT email, subscribed_at FROM subscribers ORDER BY subscribed_at DESC").fetchall()

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["email", "subscribed_at"])
    for s in subs:
        writer.writerow(s)

    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = (
        f"attachment; filename=subscribers_{datetime.now().strftime('%Y%m%d')}.csv"
    )
    return response


@router.get("/subscribers/export/excel")
async def export_subscribers_excel(user=Depends(require_auth)):
    with get_db_connection() as conn:
        subs = conn.execute("SELECT email, subscribed_at FROM subscribers ORDER BY subscribed_at DESC").fetchall()

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscribers"

    # Add headers with styling
    headers = ["Email", "Subscribed At"]
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for sub in subs:
        ws.append([sub[0], sub[1]])

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = (
        f"attachment; filename=subscribers_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    return response


@router.get("/messages", response_model=List[Message])
async def get_messages(user=Depends(require_auth)):
    with get_db_connection() as conn:
        conn.cursor().execute("SELECT id, name, email, message, sent_at FROM messages ORDER BY sent_at DESC")
        return [dict(r) for r in conn.cursor().fetchall()]


@router.delete("/messages/{id}")
async def delete_message(id: int, user=Depends(require_auth)):
    with get_db_connection() as conn:
        conn.execute("DELETE FROM messages WHERE id = ?", (id,))
        conn.commit()
    return Response(status_code=204)


# --- Public Endpoints ---


@router.get("/pages/public", dependencies=[Depends(limiter_standard)])
async def get_public_pages():
    """Get list of active pages for public use (e.g., newsletter redirect dropdown)."""
    with get_db_connection() as conn:
        pages = conn.execute(
            "SELECT id, slug, title FROM pages WHERE is_active = 1 ORDER BY created_at ASC"
        ).fetchall()
        return [{"id": p[0], "slug": p[1], "title": p[2]} for p in pages]


@router.get("/items", response_model=List[Item], dependencies=[Depends(limiter_standard)])
async def get_public_items(request: Request, page_id: Optional[int] = None):
    user = await check_auth(request)
    cache_key = f"items_{'admin' if user else 'public'}_{page_id or 'all'}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    query = "SELECT * FROM items"
    params = []
    conditions = []
    
    if page_id is not None:
        conditions.append("page_id = ?")
        params.append(page_id)
    
    if user is None:
        conditions.append("is_active = 1")
        conditions.append("(publish_on IS NULL OR publish_on <= datetime('now', 'localtime'))")
        conditions.append("(expires_on IS NULL OR expires_on >= datetime('now', 'localtime'))")
    
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY display_order ASC"

    with get_db_connection() as conn:
        rows = conn.execute(query, params).fetchall()

    items_dict = {}
    nested = []
    # Always create a new Item with a fresh children list
    for r in rows:
        item_data = dict(r)
        item_data["children"] = []
        item = Item(**item_data)
        items_dict[item.id] = item
        if item.parent_id is None:
            nested.append(item)

    for item in items_dict.values():
        if item.parent_id in items_dict:
            if item in nested:
                nested.remove(item)
            items_dict[item.parent_id].children.append(item)

    cache.set(cache_key, nested, ttl=300)
    return nested


@router.post("/click/{item_id}", dependencies=[Depends(limiter_strict)])
async def track_click(item_id: int, request: Request):
    try:
        referer = request.headers.get("referer")
        domain = urlparse(referer).netloc if referer else "(Direkt)"
        ip = request.client.host
        country = await get_country_from_ip(ip)
        with get_db_connection() as conn:
            conn.execute(
                "INSERT INTO clicks (item_id, referer, country_code) VALUES (?, ?, ?)", (item_id, domain, country)
            )
            conn.commit()
    except:
        pass
    return Response(status_code=204)


@router.post("/subscribe", dependencies=[Depends(limiter_strict)])
async def subscribe(req: SubscribeRequest):
    if not req.privacy_agreed:
        raise HTTPException(400, "Datenschutz nicht akzeptiert")
    try:
        redirect_url = None
        with get_db_connection() as conn:
            # Validate redirect_page_id if provided
            if req.redirect_page_id:
                page = conn.execute("SELECT slug FROM pages WHERE id = ? AND is_active = 1", (req.redirect_page_id,)).fetchone()
                if not page:
                    # Invalid or inactive page, ignore redirect
                    req.redirect_page_id = None
                else:
                    slug = page[0]
                    redirect_url = f"/{slug}" if slug else "/"
            
            conn.execute("INSERT INTO subscribers (email, redirect_page_id) VALUES (?, ?)", (req.email, req.redirect_page_id))
            conn.commit()
            
        return {"message": "Abonniert!", "redirect_url": redirect_url}
    except sqlite3.IntegrityError:
        return {"message": "Bereits registriert.", "redirect_url": None}


@router.post("/contact", dependencies=[Depends(limiter_strict)])
async def contact(req: ContactRequest):
    if not req.privacy_agreed:
        raise HTTPException(400, "Datenschutz nicht akzeptiert")
    with get_db_connection() as conn:
        conn.execute("INSERT INTO messages (name, email, message) VALUES (?, ?, ?)", (req.name, req.email, req.message))
        conn.commit()
    return {"message": "Gesendet!"}


# --- Tools ---


@router.get("/social/card.png")
async def social_card():
    settings = get_settings_from_db()
    img = await generate_social_card(settings)
    return StreamingResponse(img, media_type="image/png")


@router.get("/contact.vcf", dependencies=[Depends(limiter_standard)])
async def vcard():
    s = get_settings_from_db()
    vcf = f"BEGIN:VCARD\nVERSION:3.0\nFN:{s.get('title')}\nEMAIL:{s.get('social_email')}\nURL:https://{APP_DOMAIN}\nNOTE:{s.get('bio')}\nEND:VCARD"
    return Response(
        content=vcf, media_type="text/vcard", headers={"Content-Disposition": "attachment; filename=contact.vcf"}
    )


@router.get("/qrcode", dependencies=[Depends(limiter_standard)])
async def get_qr():
    try:
        # Import hier lokal, damit Fehler abgefangen werden kann
        import qrcode

        qr = qrcode.QRCode(box_size=10, border=4)
        qr.add_data(f"https://{APP_DOMAIN}")
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, "PNG")
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")
    except ImportError:
        raise HTTPException(501, "QR-Code Modul fehlt. Bitte 'pip install qrcode[pil]' ausführen.")


# --- Special Pages Management ---


@router.get("/special-pages", dependencies=[Depends(require_auth)])
async def get_special_pages_list():
    """Get all special pages for admin panel."""
    from .database import get_all_special_pages
    pages = get_all_special_pages()
    return {"pages": pages}


@router.get("/special-pages/{page_key}", dependencies=[Depends(require_auth)])
async def get_special_page_content(page_key: str):
    """Get specific special page content."""
    from .database import get_special_page, get_special_page_blocks
    page = get_special_page(page_key)
    if not page:
        raise HTTPException(404, "Seite nicht gefunden")
    
    # Get blocks if they exist
    blocks = get_special_page_blocks(page_key)
    page['blocks'] = blocks
    
    return page


@router.put("/special-pages/{page_key}", dependencies=[Depends(require_auth)])
async def update_special_page_content(page_key: str, request: Request):
    """Update special page content."""
    from .database import update_special_page
    data = await request.json()
    title = data.get("title", "")
    subtitle = data.get("subtitle", "")
    content = data.get("content", "")
    
    if not title or not content:
        raise HTTPException(400, "Titel und Inhalt sind erforderlich")
    
    update_special_page(page_key, title, subtitle, content)
    return {"message": "Seite aktualisiert"}


# Special Page Blocks endpoints
@router.get("/special-pages/{page_key}/blocks", dependencies=[Depends(require_auth)])
async def get_special_page_blocks_endpoint(page_key: str):
    """Get all blocks for a special page."""
    from .database import get_special_page_blocks
    blocks = get_special_page_blocks(page_key)
    return {"blocks": blocks}


@router.post("/special-pages/{page_key}/blocks", dependencies=[Depends(require_auth)])
async def save_special_page_blocks_endpoint(page_key: str, request: Request):
    """Save blocks for a special page."""
    from .database import save_special_page_blocks
    data = await request.json()
    blocks = data.get("blocks", [])
    
    save_special_page_blocks(page_key, blocks)
    return {"message": "Blöcke gespeichert"}


@router.put("/special-page-blocks/{block_id}", dependencies=[Depends(require_auth)])
async def update_special_page_block_endpoint(block_id: int, request: Request):
    """Update a specific block."""
    from .database import update_special_page_block
    data = await request.json()
    content = data.get("content", "")
    settings = data.get("settings", {})
    
    update_special_page_block(block_id, content, settings)
    return {"message": "Block aktualisiert"}


@router.delete("/special-page-blocks/{block_id}", dependencies=[Depends(require_auth)])
async def delete_special_page_block_endpoint(block_id: int):
    """Delete a specific block."""
    from .database import delete_special_page_block
    delete_special_page_block(block_id)
    return {"message": "Block gelöscht"}


# --- Media Kit Management ---


@router.get("/mediakit-data", dependencies=[Depends(require_auth)])
async def get_mediakit_admin_data():
    """Get all media kit data for admin panel."""
    data = get_mediakit_data()
    return {"data": data}


@router.put("/mediakit-data", dependencies=[Depends(require_auth)])
async def update_mediakit_admin_data(request: Request):
    """Update media kit data."""
    data = await request.json()
    section = data.get("section", "")
    key = data.get("key", "")
    value = data.get("value", "")
    display_order = data.get("display_order", 0)
    
    if not section or not key:
        raise HTTPException(400, "Section und Key sind erforderlich")
    
    update_mediakit_data(section, key, value, display_order)
    return {"message": "Media Kit Daten aktualisiert"}


@router.post("/mediakit-data/batch", dependencies=[Depends(require_auth)])
async def update_mediakit_batch(request: Request):
    """Batch update media kit data."""
    data = await request.json()
    updates = data.get("updates", [])
    
    if not updates:
        raise HTTPException(400, "Keine Updates bereitgestellt")
    
    for update in updates:
        section = update.get("section", "")
        key = update.get("key", "")
        value = update.get("value", "")
        display_order = update.get("display_order", 0)
        
        if section and key:
            update_mediakit_data(section, key, value, display_order)
    
    return {"message": "Alle Media Kit Daten aktualisiert"}


@router.delete("/mediakit-data", dependencies=[Depends(require_auth)])
async def delete_mediakit_admin_data(request: Request):
    """Delete media kit entry."""
    data = await request.json()
    section = data.get("section", "")
    key = data.get("key", "")
    
    if not section or not key:
        raise HTTPException(400, "Section und Key sind erforderlich")
    
    delete_mediakit_entry(section, key)
    return {"message": "Eintrag gelöscht"}


@router.post("/mediakit/refresh-social-stats", dependencies=[Depends(require_auth)])
async def refresh_social_stats(request: Request):
    """Fetch fresh social media statistics from Profile tab handles and update cache."""
    # Get social handles from Profile tab (settings) instead of mediakit_data
    settings = get_settings_from_db()
    
    config = {
        'instagram_handle': settings.get('social_instagram', ''),
        'tiktok_handle': settings.get('social_tiktok', ''),
        'youtube_handle': settings.get('social_youtube', ''),
        'twitch_handle': settings.get('social_twitch', ''),
        'x_handle': settings.get('social_x', ''),
    }
    
    # Remove empty handles
    config = {k: v for k, v in config.items() if v}
    
    if not config:
        raise HTTPException(400, "Keine Social Media Handles konfiguriert. Bitte zuerst im Profil-Tab Social Media Handles eingeben und speichern.")
    
    # Fetch stats
    stats_service = get_stats_service()
    results = await stats_service.fetch_all_stats(config)
    
    # Check if we got any data
    if not results.get('platforms'):
        # If scraping failed, provide helpful error message
        error_msg = "Konnte keine Daten abrufen. "
        if results.get('errors'):
            error_msg += " Mögliche Gründe: " + ", ".join(results['errors'][:2])
        raise HTTPException(500, error_msg)
    
    # Save to cache with full analytics data
    for platform, stats in results.get('platforms', {}).items():
        save_social_stats_cache(platform, stats['username'], json.dumps(stats))
    
    # Update mediakit_data with new stats for backward compatibility
    if 'instagram' in results['platforms']:
        ig_stats = results['platforms']['instagram']
        followers = stats_service.format_number(ig_stats.get('followers', 0))
        update_mediakit_data('platforms', 'instagram_followers', followers, 1)
        update_mediakit_data('platforms', 'instagram_handle', f"@{ig_stats['username']}", 1)
    
    if 'tiktok' in results['platforms']:
        tt_stats = results['platforms']['tiktok']
        followers = stats_service.format_number(tt_stats.get('followers', 0))
        update_mediakit_data('platforms', 'tiktok_followers', followers, 2)
        update_mediakit_data('platforms', 'tiktok_handle', f"@{tt_stats['username']}", 2)
    
    # Update total
    total = stats_service.format_number(results.get('total_followers', 0))
    update_mediakit_data('analytics', 'total_followers', total, 0)
    
    # Store last update timestamp
    update_mediakit_data('analytics', 'last_updated', datetime.now().strftime('%d.%m.%Y'), 99)
    
    return {
        "message": "Social Media Statistiken erfolgreich aktualisiert",
        "data": results,
        "total_followers": total,
        "total_followers_raw": results.get('total_followers', 0),
        "platforms_updated": list(results.get('platforms', {}).keys())
    }


@router.get("/mediakit/social-stats-cache")
async def get_cached_social_stats():
    """Get cached social media statistics (public endpoint for frontend)."""
    cache = get_social_stats_cache()
    return {"data": cache}


@router.get("/mediakit/follower-summary")
async def get_follower_summary():
    """Get follower summary for platforms with 1000+ followers (public endpoint)."""
    cache = get_social_stats_cache()
    stats_service = get_stats_service()
    
    qualified_platforms = []
    total_qualified_followers = 0
    
    for platform, data in cache.items():
        stats = data.get('data', {})
        followers = stats.get('followers', 0)
        
        # Only include platforms with 1000+ followers
        if followers >= 1000:
            qualified_platforms.append({
                'platform': platform,
                'username': data.get('username', ''),
                'followers': followers,
                'followers_formatted': stats_service.format_number(followers),
                'verified': stats.get('verified', False),
                'fetched_at': data.get('fetched_at', '')
            })
            total_qualified_followers += followers
    
    return {
        "total_followers": total_qualified_followers,
        "total_followers_formatted": stats_service.format_number(total_qualified_followers),
        "platforms": qualified_platforms,
        "platform_count": len(qualified_platforms)
    }


@router.get("/mediakit/analytics/{platform}")
async def get_platform_analytics(platform: str):
    """Get detailed analytics for a specific platform (public endpoint)."""
    cache = get_social_stats_cache(platform)
    
    # Check if we got any data back for this platform
    if not cache:
        raise HTTPException(404, f"Keine Daten für Plattform '{platform}' gefunden")
    
    # The filtered cache should have the platform as a key
    if platform not in cache:
        raise HTTPException(404, f"Keine Daten für Plattform '{platform}' gefunden")
    
    data = cache[platform]
    stats = data.get('data', {})
    
    # Build platform-specific analytics response
    analytics = {
        'platform': platform,
        'username': data.get('username', ''),
        'fetched_at': data.get('fetched_at', ''),
        'metrics': {}
    }
    
    # Common metrics
    if 'followers' in stats:
        analytics['metrics']['followers'] = stats['followers']
    if 'following' in stats:
        analytics['metrics']['following'] = stats['following']
    
    # Platform-specific metrics
    if platform == 'instagram':
        analytics['metrics']['posts'] = stats.get('posts', 0)
        analytics['metrics']['engagement_rate'] = stats.get('engagement_rate')
        analytics['metrics']['verified'] = stats.get('verified', False)
    elif platform == 'tiktok':
        analytics['metrics']['likes'] = stats.get('likes', 0)
        analytics['metrics']['videos'] = stats.get('videos', 0)
    elif platform == 'youtube':
        analytics['metrics']['subscribers'] = stats.get('subscribers', 0)
        analytics['metrics']['videos'] = stats.get('videos', 0)
        analytics['metrics']['views'] = stats.get('views', 0)
    
    return analytics


@router.post("/mediakit/refresh-instagram-api", dependencies=[Depends(require_auth)])
async def refresh_instagram_api_stats():
    """
    Fetch fresh Instagram statistics using Meta Graph API (not scraping).
    Requires .env.social file with Instagram API credentials.
    """
    from dotenv import load_dotenv
    from .instagram_fetcher import get_instagram_fetcher_from_env
    
    # Constants
    TOKEN_PREVIEW_LENGTH = 20  # Number of chars to show from token for security
    
    # Load .env.social if it exists
    social_env_path = Path(__file__).parent.parent / '.env.social'
    if social_env_path.exists():
        load_dotenv(social_env_path)
    else:
        raise HTTPException(
            400, 
            "Keine .env.social Datei gefunden. Bitte Instagram API Credentials konfigurieren."
        )
    
    # Get Instagram fetcher
    fetcher = get_instagram_fetcher_from_env()
    if not fetcher:
        raise HTTPException(
            400,
            "Instagram API nicht konfiguriert. Prüfe INSTAGRAM_ACCESS_TOKEN und INSTAGRAM_USERNAME in .env.social"
        )
    
    try:
        # Fetch stats and refresh token
        stats, new_token = await fetcher.fetch_and_refresh_token()
        
        if not stats:
            raise HTTPException(500, "Konnte Instagram Daten nicht abrufen")
        
        # Save to cache
        save_social_stats_cache(
            platform='instagram',
            username=stats['profile']['username'],
            stats_data=json.dumps(stats)
        )
        
        # Update mediakit_data for backward compatibility
        stats_service = get_stats_service()
        followers = stats_service.format_number(stats['stats']['followers'])
        update_mediakit_data('platforms', 'instagram_followers', followers, 1)
        update_mediakit_data('platforms', 'instagram_handle', f"@{stats['profile']['username']}", 1)
        
        # Update analytics
        update_mediakit_data('analytics', 'last_updated', datetime.now().strftime('%d.%m.%Y'), 99)
        
        response_data = {
            "message": "Instagram Statistiken erfolgreich über API aktualisiert",
            "data": {
                "username": stats['profile']['username'],
                "followers": stats['stats']['followers'],
                "followers_formatted": followers,
                "posts": stats['stats']['posts'],
                "reach_daily": stats['stats']['reach_daily'],
                "impressions_daily": stats['stats']['impressions_daily'],
                "profile_views": stats['stats']['profile_views'],
            },
            "source": "Meta Graph API",
            "updated_at": stats['meta']['updated_at']
        }
        
        # Include token refresh info if token was refreshed
        if new_token:
            response_data["token_refreshed"] = True
            response_data["warning"] = "Access Token wurde erneuert. Bitte GitHub Secret 'INSTAGRAM_SECRET' aktualisieren!"
            response_data["new_token"] = new_token[:TOKEN_PREVIEW_LENGTH] + "..."  # Show only first chars for security
        
        return response_data
        
    except Exception as e:
        logging.error(f"Error fetching Instagram API stats: {e}")
        raise HTTPException(500, f"Fehler beim Abrufen der Instagram Daten: {str(e)}")


# Media Kit Settings Endpoints
@router.get("/mediakit/settings", dependencies=[Depends(require_auth)])
async def get_mediakit_settings_endpoint():
    """Get all media kit settings."""
    settings = get_all_mediakit_settings()
    return {"settings": settings}


@router.put("/mediakit/settings", dependencies=[Depends(require_auth)])
async def update_mediakit_settings_endpoint(request: Request):
    """Update media kit settings."""
    data = await request.json()
    
    for key, value in data.items():
        update_mediakit_setting(key, str(value) if value is not None else "")
    
    return {"message": "Media Kit Einstellungen aktualisiert"}


# Media Kit Views Tracking Endpoints
@router.post("/mediakit/track-view", dependencies=[Depends(limiter_standard)])
async def track_mediakit_view_endpoint(request: Request):
    """Track a media kit view (public endpoint)."""
    data = await request.json()
    
    # Get viewer info
    viewer_email = data.get("email")
    viewer_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")
    
    # Get country from IP (async)
    viewer_country = None
    if viewer_ip:
        viewer_country = await get_country_from_ip(viewer_ip)
    
    # Track the view
    track_mediakit_view(
        viewer_email=viewer_email,
        viewer_ip=viewer_ip,
        viewer_country=viewer_country,
        user_agent=user_agent
    )
    
    return {"message": "View tracked successfully"}


@router.get("/mediakit/views", dependencies=[Depends(require_auth)])
async def get_mediakit_views_endpoint(limit: int = 100):
    """Get recent media kit views."""
    views = get_mediakit_views(limit)
    return {"views": views}


@router.get("/mediakit/views/stats", dependencies=[Depends(require_auth)])
async def get_mediakit_views_stats_endpoint():
    """Get media kit views statistics."""
    stats = get_mediakit_views_stats()
    return stats


# Media Kit Access Request Endpoints
@router.post("/mediakit/request-access", dependencies=[Depends(limiter_strict)])
async def request_mediakit_access(request: Request):
    """Request access to view media kit (for gated access)."""
    data = await request.json()
    
    email = data.get("email", "").strip()
    name = data.get("name", "").strip()
    company = data.get("company", "").strip()
    message = data.get("message", "").strip()
    
    if not email:
        raise HTTPException(400, "Email ist erforderlich")
    
    # Validate email format
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        raise HTTPException(400, "Ungültige E-Mail-Adresse")
    
    # Get IP address
    ip_address = request.client.host if request.client else None
    
    # Create access request
    request_id = create_access_request(
        email=email,
        name=name,
        company=company,
        message=message,
        ip_address=ip_address
    )
    
    return {
        "message": "Zugriffsanfrage wurde gesendet. Sie erhalten eine E-Mail, sobald Ihre Anfrage genehmigt wurde.",
        "request_id": request_id
    }


@router.get("/mediakit/access-requests", dependencies=[Depends(require_auth)])
async def get_access_requests_endpoint(status: Optional[str] = None):
    """Get access requests."""
    requests = get_access_requests(status)
    return {"requests": requests}


@router.put("/mediakit/access-requests/{request_id}", dependencies=[Depends(require_auth)])
async def update_access_request_endpoint(request_id: int, request: Request):
    """Update access request status."""
    data = await request.json()
    status = data.get("status", "").strip()
    
    if status not in ["approved", "rejected", "pending"]:
        raise HTTPException(400, "Ungültiger Status")
    
    update_access_request_status(request_id, status)
    
    return {"message": f"Zugriffsanfrage wurde {status}"}


@router.get("/mediakit/check-access")
async def check_mediakit_access(email: str):
    """Check if an email has approved access."""
    has_access = check_access_approved(email)
    return {"has_access": has_access}


# Media Kit PDF Export Endpoint
@router.get("/mediakit/export/pdf", dependencies=[Depends(require_auth)])
async def export_mediakit_pdf():
    """Export media kit as PDF."""
    try:
        # We'll use a library to generate PDF from HTML
        # For now, return a placeholder response
        # In production, you'd use something like weasyprint or pdfkit
        
        raise HTTPException(501, "PDF Export wird in Kürze verfügbar sein. Bitte verwenden Sie vorerst die Browser-Druckfunktion (Strg+P).")
        
    except Exception as e:
        logging.error(f"Error exporting PDF: {e}")
        raise HTTPException(500, "Fehler beim Exportieren des PDFs")


# New Media Kit Blocks API (Block-based system)
@router.get("/mediakit/blocks", dependencies=[Depends(require_auth)])
async def get_all_mediakit_blocks():
    """Get all media kit blocks for admin."""
    try:
        blocks = get_mediakit_blocks()
        return {"blocks": blocks}
    except Exception as e:
        logger.error(f"Error getting media kit blocks: {e}")
        return JSONResponse(status_code=500, content={"detail": str(e)})


@router.post("/mediakit/blocks", dependencies=[Depends(require_auth)])
async def create_new_mediakit_block(request: Request):
    """Create a new media kit block."""
    try:
        data = await request.json()
        block_type = data.get('block_type')
        title = data.get('title')
        content = data.get('content')
        settings = data.get('settings', {})
        position = data.get('position')
        
        if not block_type:
            return JSONResponse(status_code=400, content={"detail": "block_type is required"})
        
        block_id = create_mediakit_block(block_type, title, content, settings, position)
        return {"success": True, "block_id": block_id}
    except Exception as e:
        logger.error(f"Error creating media kit block: {e}")
        return JSONResponse(status_code=500, content={"detail": str(e)})


@router.put("/mediakit/blocks/{block_id}", dependencies=[Depends(require_auth)])
async def update_existing_mediakit_block(block_id: int, request: Request):
    """Update a media kit block."""
    try:
        data = await request.json()
        title = data.get('title')
        content = data.get('content')
        settings = data.get('settings')
        is_visible = data.get('is_visible')
        
        update_mediakit_block(block_id, title, content, settings, is_visible)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error updating media kit block: {e}")
        return JSONResponse(status_code=500, content={"detail": str(e)})


@router.delete("/mediakit/blocks/{block_id}", dependencies=[Depends(require_auth)])
async def delete_existing_mediakit_block(block_id: int):
    """Delete a media kit block."""
    try:
        delete_mediakit_block(block_id)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error deleting media kit block: {e}")
        return JSONResponse(status_code=500, content={"detail": str(e)})


@router.post("/mediakit/blocks/reorder", dependencies=[Depends(require_auth)])
async def reorder_blocks(request: Request):
    """Reorder media kit blocks."""
    try:
        data = await request.json()
        block_positions = data.get('blocks', [])
        reorder_mediakit_blocks(block_positions)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error reordering media kit blocks: {e}")
        return JSONResponse(status_code=500, content={"detail": str(e)})
