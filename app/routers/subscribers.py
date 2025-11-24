"""
Subscribers Router
Handles subscriber management, exports, and contact messages.
"""
import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from typing import List

from ..models import Subscriber, Message
from ..database import get_db_connection
from ..auth_unified import require_auth

router = APIRouter()


@router.get("", response_model=List[Subscriber])
async def get_subscribers(user=Depends(require_auth)):
    """Get all subscribers."""
    with get_db_connection() as conn:
        rows = conn.execute("SELECT id, email, subscribed_at, redirect_page_id FROM subscribers ORDER BY subscribed_at DESC").fetchall()
        return [dict(r) for r in rows]


@router.delete("/{id}")
async def delete_subscriber(id: int, user=Depends(require_auth)):
    """Delete a subscriber."""
    with get_db_connection() as conn:
        conn.execute("DELETE FROM subscribers WHERE id = ?", (id,))
        conn.commit()
    return Response(status_code=204)


@router.get("/export")
async def export_subscribers(user=Depends(require_auth)):
    """Export subscribers as CSV."""
    with get_db_connection() as conn:
        subs = conn.execute("SELECT email, subscribed_at FROM subscribers ORDER BY subscribed_at DESC").fetchall()

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["email", "subscribed_at"])
    for s in subs:
        writer.writerow(s)

    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = (
        f"attachment; filename=subscribers_{datetime.now().strftime('%Y%m%d')}.csv"
    )
    return response


@router.get("/export/excel")
async def export_subscribers_excel(user=Depends(require_auth)):
    """Export subscribers as Excel file."""
    with get_db_connection() as conn:
        subs = conn.execute("SELECT email, subscribed_at FROM subscribers ORDER BY subscribed_at DESC").fetchall()

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscribers"

    # Add headers with styling
    headers = ["Email", "Subscribed At"]
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for sub in subs:
        ws.append([sub[0], sub[1]])

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = (
        f"attachment; filename=subscribers_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    return response


@router.get("/messages", response_model=List[Message])
async def get_messages(user=Depends(require_auth)):
    """Get all contact messages."""
    with get_db_connection() as conn:
        rows = conn.execute("SELECT id, name, email, message, sent_at FROM messages ORDER BY sent_at DESC").fetchall()
        return [dict(r) for r in rows]


@router.delete("/messages/{id}")
async def delete_message(id: int, user=Depends(require_auth)):
    """Delete a contact message."""
    with get_db_connection() as conn:
        conn.execute("DELETE FROM messages WHERE id = ?", (id,))
        conn.commit()
    return Response(status_code=204)
